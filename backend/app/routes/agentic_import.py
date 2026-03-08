"""
Agentic file import with AI-driven format detection, balance extraction,
code generation, and reusable format profiles.

Endpoints:
  POST /upload       — Upload file, compute fingerprint, check for profile
  POST /analyze      — Run AI analysis (with optional clarification)
  POST /approve      — Save profile + import transactions
  GET  /profiles     — List user's saved format profiles
"""
import csv
import io
import json
import logging
import os
import tempfile
import uuid
from datetime import datetime, date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.db_helpers import get_user_id
from app.models import Account, FormatProfile, Transaction
from app.services.format_fingerprint import compute_fingerprint
from app.services.import_script_generator import (
    check_clarity,
    generate_script_with_retry,
)
from app.services.script_sandbox import execute_script

logger = logging.getLogger(__name__)

router = APIRouter()

# ---------------------------------------------------------------------------
# In-memory store for pending import sessions (keyed by import_id).
# Production deployments should back this with Redis or the database.
# ---------------------------------------------------------------------------
_pending_sessions: Dict[str, Dict[str, Any]] = {}

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class UploadResponse(BaseModel):
    import_id: str
    status: str  # "profile_matched" | "needs_analysis"
    fingerprint: str
    headers: List[str]
    sample_rows: Optional[List[List[str]]] = None
    profile_label: Optional[str] = None
    # For profile_matched: preview data is included directly
    mapping_summary: Optional[str] = None
    transformation_description: Optional[str] = None
    balance_column: Optional[str] = None
    sample_transactions: Optional[List[Dict[str, Any]]] = None
    total_rows: Optional[int] = None


class AnalyzeRequest(BaseModel):
    import_id: str
    clarification_response: Optional[str] = None


class AnalyzeResponse(BaseModel):
    status: str  # "preview_ready" | "needs_clarification" | "failed"
    question: Optional[str] = None
    mapping_summary: Optional[str] = None
    transformation_description: Optional[str] = None
    balance_column: Optional[str] = None
    sample_transactions: Optional[List[Dict[str, Any]]] = None
    total_rows: Optional[int] = None
    error: Optional[str] = None


class ApproveRequest(BaseModel):
    import_id: str
    account_id: str


class FailedRow(BaseModel):
    row_number: int
    reason: str


class ApproveResponse(BaseModel):
    success: bool
    total_rows: int
    imported: int
    duplicates_skipped: int
    failed_rows: List[FailedRow]
    balance_anchors_detected: bool
    error: Optional[str] = None


class ProfileItem(BaseModel):
    id: str
    fingerprint: str
    label: str
    created_at: str


class ProfilesResponse(BaseModel):
    profiles: List[ProfileItem]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_headers_and_rows(file_path: str, filename: str):
    """Extract column headers and data rows from CSV or XLSX."""
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".xls":
        raise HTTPException(
            status_code=400,
            detail="XLS files are not supported. Please re-export your file as XLSX or CSV.",
        )

    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        candidate_sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row and any(v is not None for v in first_row):
                row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
                if row_count > 0:
                    candidate_sheets.append((name, first_row, row_count))

        if not candidate_sheets:
            raise HTTPException(status_code=400, detail="No sheets with tabular data found in the XLSX file.")

        # Auto-select if only one candidate; otherwise pick the largest
        selected = max(candidate_sheets, key=lambda x: x[2])
        ws = wb[selected[0]]

        headers = [str(c) if c is not None else "" for c in selected[1]]
        rows: List[List[str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])

        wb.close()
        return headers, rows

    # Default: CSV
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            content = f.read()

    # Detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(content[:4096])
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise HTTPException(status_code=400, detail="CSV file is empty.")

    headers = all_rows[0]
    rows = [r for r in all_rows[1:] if any(cell.strip() for cell in r)]
    return headers, rows


def _compute_daily_balances(transactions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Extract closing balance per day from parsed transactions."""
    daily: Dict[str, float] = {}
    for tx in transactions:
        bal = tx.get("balance")
        if bal is None:
            continue
        try:
            bal_float = float(bal)
        except (ValueError, TypeError):
            continue
        d = tx.get("date", "")[:10]
        if d:
            daily[d] = bal_float  # last row per day wins (list is sorted by date asc)
    return [{"date": d, "balance": b} for d, b in sorted(daily.items())]


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/upload", response_model=UploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    account_id: str = Form(...),
    user_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Upload a bank file, compute fingerprint, and check for a matching profile."""
    resolved_user_id = get_user_id(user_id)

    # Validate account
    account = db.query(Account).filter(
        Account.id == account_id, Account.user_id == resolved_user_id
    ).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    # Validate file type
    filename = file.filename or "upload.csv"
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".csv", ".xlsx"):
        if ext == ".xls":
            raise HTTPException(
                status_code=400,
                detail="XLS files are not supported. Please re-export your file as XLSX or CSV.",
            )
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported.")

    # Save to temp file
    suffix = ext
    fd, temp_path = tempfile.mkstemp(suffix=suffix)
    try:
        content = await file.read()
        with os.fdopen(fd, "wb") as f:
            f.write(content)
    except Exception:
        os.close(fd)
        raise

    try:
        headers, rows = _extract_headers_and_rows(temp_path, filename)
    except HTTPException:
        os.unlink(temp_path)
        raise
    except Exception as e:
        os.unlink(temp_path)
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    fingerprint = compute_fingerprint(headers)
    import_id = str(uuid.uuid4())

    # Check for existing profile
    profile = db.query(FormatProfile).filter(
        FormatProfile.user_id == resolved_user_id,
        FormatProfile.fingerprint == fingerprint,
    ).first()

    sample_rows = rows[:10]

    if profile and profile.script:
        # Run stored script
        logger.info(f"Found matching profile '{profile.label}' for fingerprint {fingerprint[:12]}…")
        success, result, error = execute_script(profile.script, temp_path)

        if success and result:
            _pending_sessions[import_id] = {
                "user_id": resolved_user_id,
                "account_id": account_id,
                "file_path": temp_path,
                "filename": filename,
                "fingerprint": fingerprint,
                "headers": headers,
                "rows": rows,
                "transactions": result,
                "script": profile.script,
                "profile_id": str(profile.id),
                "is_new_profile": False,
                "mapping_summary": profile.label,
                "transformation_description": "",
                "balance_column": None,
            }

            balance_col = None
            if result and result[0].get("balance") is not None:
                balance_col = "balance"

            return UploadResponse(
                import_id=import_id,
                status="profile_matched",
                fingerprint=fingerprint,
                headers=headers,
                sample_rows=sample_rows,
                profile_label=profile.label,
                mapping_summary=f"Recognised format — using saved profile: {profile.label}",
                transformation_description="",
                balance_column=balance_col,
                sample_transactions=result[:10],
                total_rows=len(result),
            )
        else:
            logger.warning(f"Stored script failed for profile {profile.id}: {error}")

    # No matching profile (or stored script failed) — needs analysis
    _pending_sessions[import_id] = {
        "user_id": resolved_user_id,
        "account_id": account_id,
        "file_path": temp_path,
        "filename": filename,
        "fingerprint": fingerprint,
        "headers": headers,
        "rows": rows,
        "transactions": None,
        "script": None,
        "profile_id": None,
        "is_new_profile": True,
        "clarification_history": [],
    }

    return UploadResponse(
        import_id=import_id,
        status="needs_analysis",
        fingerprint=fingerprint,
        headers=headers,
        sample_rows=sample_rows,
    )


@router.post("/analyze", response_model=AnalyzeResponse)
def analyze_file(
    request: AnalyzeRequest,
    db: Session = Depends(get_db),
):
    """Run AI analysis on the uploaded file, optionally incorporating a clarification response."""
    session = _pending_sessions.get(request.import_id)
    if not session:
        raise HTTPException(status_code=404, detail="Import session not found or expired")

    resolved_user_id = get_user_id(None)
    if session["user_id"] != resolved_user_id:
        raise HTTPException(status_code=403, detail="Not authorised")

    headers = session["headers"]
    rows = session["rows"]
    sample_rows = rows[:10]

    # Build clarification context
    clarification_history = session.get("clarification_history", [])
    clarification_context = None

    if request.clarification_response:
        clarification_history.append({
            "answer": request.clarification_response,
        })
        session["clarification_history"] = clarification_history
        clarification_context = request.clarification_response

    # Check if structure is clear
    clarity_result = check_clarity(headers, sample_rows, clarification_history if clarification_history else None)

    if clarity_result.get("status") == "needs_clarification":
        return AnalyzeResponse(
            status="needs_clarification",
            question=clarity_result.get("question", "Could you describe the format of your bank file?"),
        )

    # Structure is clear — generate and execute script
    file_path = session["file_path"]
    combined_context = ""
    for entry in clarification_history:
        combined_context += entry.get("answer", "") + "\n"
    if clarity_result.get("transformation_description"):
        combined_context += clarity_result["transformation_description"]

    success, script, transactions, error = generate_script_with_retry(
        headers, sample_rows, file_path,
        clarification_context=combined_context.strip() or None,
    )

    if not success:
        return AnalyzeResponse(
            status="failed",
            error=error or "We couldn't process this file. Try re-exporting it from your bank or contact support.",
        )

    session["transactions"] = transactions
    session["script"] = script
    session["mapping_summary"] = clarity_result.get("mapping_summary", "")
    session["transformation_description"] = clarity_result.get("transformation_description", "")
    session["balance_column"] = clarity_result.get("balance_column")

    balance_col = None
    if transactions and transactions[0].get("balance") is not None:
        balance_col = clarity_result.get("balance_column") or "balance"

    return AnalyzeResponse(
        status="preview_ready",
        mapping_summary=clarity_result.get("mapping_summary", ""),
        transformation_description=clarity_result.get("transformation_description", ""),
        balance_column=balance_col,
        sample_transactions=transactions[:10] if transactions else [],
        total_rows=len(transactions) if transactions else 0,
    )


@router.post("/approve", response_model=ApproveResponse)
def approve_import(
    request: ApproveRequest,
    db: Session = Depends(get_db),
):
    """Approve the import: save format profile, run duplicate detection, and import."""
    session = _pending_sessions.get(request.import_id)
    if not session:
        raise HTTPException(status_code=404, detail="Import session not found or expired")

    resolved_user_id = get_user_id(None)
    if session["user_id"] != resolved_user_id:
        raise HTTPException(status_code=403, detail="Not authorised")

    transactions = session.get("transactions")
    if not transactions:
        raise HTTPException(status_code=400, detail="No parsed transactions available")

    account_id = request.account_id
    account = db.query(Account).filter(
        Account.id == account_id, Account.user_id == resolved_user_id
    ).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    # 1. Save format profile (if new)
    if session.get("is_new_profile") and session.get("script"):
        fingerprint = session["fingerprint"]
        headers = session["headers"]
        col_count = len(headers)
        filename = session.get("filename", "unknown")
        ext = os.path.splitext(filename)[1].upper().lstrip(".")
        label = f"{ext} — {col_count} columns"

        existing = db.query(FormatProfile).filter(
            FormatProfile.user_id == resolved_user_id,
            FormatProfile.fingerprint == fingerprint,
        ).first()

        if not existing:
            profile = FormatProfile(
                user_id=resolved_user_id,
                fingerprint=fingerprint,
                script=session["script"],
                label=label,
            )
            db.add(profile)
            db.commit()
            logger.info(f"Saved new format profile '{label}' for user {resolved_user_id}")

    # 2. Duplicate detection (date + amount + description, across all user accounts)
    duplicates_skipped = 0
    failed_rows: List[FailedRow] = []
    to_import: List[Dict[str, Any]] = []
    balance_anchors_detected = False

    for idx, tx in enumerate(transactions):
        row_num = idx + 1
        try:
            tx_date_str = tx.get("date", "")[:10]
            tx_amount = float(tx.get("amount", 0))
            tx_desc = (tx.get("description") or "").strip()

            if not tx_date_str or not tx_desc:
                failed_rows.append(FailedRow(row_number=row_num, reason="Missing date or description"))
                continue

            if tx.get("balance") is not None:
                balance_anchors_detected = True

            # Check duplicates across all user accounts
            from sqlalchemy import func, cast, Date as SqlDate
            tx_date = datetime.strptime(tx_date_str, "%Y-%m-%d").date()

            existing = db.query(Transaction).filter(
                Transaction.user_id == resolved_user_id,
                cast(Transaction.booked_at, SqlDate) == tx_date,
                Transaction.amount == tx_amount,
                func.lower(func.trim(Transaction.description)) == tx_desc.lower().strip(),
            ).first()

            if existing:
                duplicates_skipped += 1
                continue

            to_import.append(tx)

        except Exception as e:
            failed_rows.append(FailedRow(row_number=row_num, reason=str(e)))

    # 3. Import via backend transaction import pipeline
    imported = 0
    if to_import:
        from app.routes.transaction_import import TransactionImportRequest, TransactionImportItem, import_transactions
        from decimal import Decimal

        items = []
        for tx in to_import:
            amount = float(tx.get("amount", 0))
            fee = float(tx.get("fee") or 0)
            net = amount - abs(fee) if fee else amount
            tx_type = "credit" if net >= 0 else "debit"

            items.append(TransactionImportItem(
                account_id=uuid.UUID(account_id),
                amount=Decimal(str(round(net, 2))),
                description=tx.get("description") or "",
                merchant=tx.get("merchant"),
                booked_at=datetime.fromisoformat(tx["date"][:10]),
                transaction_type=tx_type,
                currency=tx.get("currency") or account.currency or "EUR",
                external_id=None,
            ))

        # Compute daily balances from parsed transactions
        daily_balances_raw = _compute_daily_balances(transactions)
        from app.schemas import DailyBalanceImport
        daily_bal_items = [DailyBalanceImport(date=d["date"], balance=d["balance"]) for d in daily_balances_raw] if daily_balances_raw else None

        import_req = TransactionImportRequest(
            transactions=items,
            user_id=resolved_user_id,
            sync_exchange_rates=True,
            update_functional_amounts=True,
            calculate_balances=True,
            detect_subscriptions=True,
            daily_balances=daily_bal_items,
        )

        try:
            result = import_transactions(import_req, db)
            imported = result.transactions_inserted
        except HTTPException as e:
            logger.error(f"Import failed: {e.detail}")
            return ApproveResponse(
                success=False,
                total_rows=len(transactions),
                imported=0,
                duplicates_skipped=duplicates_skipped,
                failed_rows=failed_rows,
                balance_anchors_detected=balance_anchors_detected,
                error=str(e.detail),
            )

    # 4. Cleanup
    try:
        file_path = session.get("file_path")
        if file_path and os.path.exists(file_path):
            os.unlink(file_path)
    except OSError:
        pass
    _pending_sessions.pop(request.import_id, None)

    return ApproveResponse(
        success=True,
        total_rows=len(transactions),
        imported=imported,
        duplicates_skipped=duplicates_skipped,
        failed_rows=failed_rows,
        balance_anchors_detected=balance_anchors_detected,
    )


@router.get("/profiles", response_model=ProfilesResponse)
def list_profiles(
    user_id: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """List user's saved format profiles."""
    resolved_user_id = get_user_id(user_id)

    profiles = db.query(FormatProfile).filter(
        FormatProfile.user_id == resolved_user_id
    ).order_by(FormatProfile.created_at.desc()).all()

    return ProfilesResponse(
        profiles=[
            ProfileItem(
                id=str(p.id),
                fingerprint=p.fingerprint,
                label=p.label or "",
                created_at=p.created_at.isoformat() if p.created_at else "",
            )
            for p in profiles
        ]
    )
